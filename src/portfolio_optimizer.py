from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Tuple

import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from scipy.optimize import minimize


BASE_DIR = Path(__file__).resolve().parents[1]
OUTPUT_DIR = BASE_DIR / "outputs"
OUTPUT_FILE = OUTPUT_DIR / "paul_asset_allocation_model.xlsx"
JPM_MATRIX_FILE = BASE_DIR / "inputs" / "matrix-usd.xlsx"
CASH_TARGET = 0.01
CATEGORIES = ["Equity", "Fixed Income", "Alternatives", "Cash"]
MODERATE_VOL_TARGET = 0.130
VOLATILITY_PERCENTILES: Dict[str, float] = {
    "Conservative": 25.0,
    "Balanced": 40.0,
    "Moderate": 55.0,
    "Growth": 70.0,
    "Aggressive Growth": 85.0,
}
TARGET_VOLATILITY_RULES: Dict[str, Tuple[float, float]] = {
    "Conservative": (-0.053, 0.005),
    "Balanced": (-0.025, 0.005),
    "Moderate": (0.000, 0.003),
    "Growth": (0.028, 0.005),
    "Aggressive Growth": (0.055, 0.005),
}
CATEGORY_BOUNDS: Dict[str, Dict[str, Tuple[float, float]]] = {
    "Conservative": {"Equity": (0.30, 0.50), "Fixed Income": (0.55, 0.75), "Alternatives": (0.00, 0.00), "Cash": (CASH_TARGET, CASH_TARGET)},
    "Balanced": {"Equity": (0.55, 0.65), "Fixed Income": (0.35, 0.55), "Alternatives": (0.00, 0.00), "Cash": (CASH_TARGET, CASH_TARGET)},
    "Moderate": {"Equity": (0.65, 0.75), "Fixed Income": (0.25, 0.40), "Alternatives": (0.00, 0.00), "Cash": (CASH_TARGET, CASH_TARGET)},
    "Growth": {"Equity": (0.75, 0.85), "Fixed Income": (0.10, 0.25), "Alternatives": (0.00, 0.00), "Cash": (CASH_TARGET, CASH_TARGET)},
    "Aggressive Growth": {"Equity": (0.90, 1.00), "Fixed Income": (0.00, 0.12), "Alternatives": (0.00, 0.00), "Cash": (CASH_TARGET, CASH_TARGET)},
}


@dataclass(frozen=True)
class Asset:
    name: str
    category: str
    source_names: Tuple[str, ...]
    min_weight: float
    max_weight: float
    source_note: str


ASSETS: List[Asset] = [
    Asset("U.S. Large Cap", "Equity", ("U.S. Large Cap",), 0.20, 0.45, "U.S. Large Cap"),
    Asset("U.S. Mid Cap", "Equity", ("U.S. Mid Cap",), 0.00, 0.20, "U.S. Mid Cap"),
    Asset("U.S. Small Cap", "Equity", ("U.S. Small Cap",), 0.00, 0.05, "U.S. Small Cap"),
    Asset("International Developed Equity", "Equity", ("EAFE Equity",), 0.10, 0.20, "EAFE Equity"),
    Asset("Emerging Markets Equity", "Equity", ("Emerging Markets Equity",), 0.00, 0.10, "Emerging Markets Equity"),
    Asset("U.S. REITs", "Alternatives", ("U.S. REITs",), 0.00, 0.00, "U.S. REITs"),
    Asset("Commodities", "Alternatives", ("Commodities",), 0.00, 0.00, "Commodities"),
    Asset("Income U.S. - U.S. Treasury", "Fixed Income", ("U.S. Intermediate Treasuries", "U.S. Long Treasuries", "TIPS"), 0.00, 0.45, "Equal-weight U.S. Intermediate Treasuries, U.S. Long Treasuries, and TIPS"),
    Asset("Income U.S. Government Related", "Fixed Income", ("U.S. Aggregate Bonds", "U.S. Muni 1-15 Yr Blend"), 0.00, 0.35, "Equal-weight U.S. Aggregate Bonds and U.S. Muni 1-15 Yr Blend"),
    Asset("Income U.S. Corporate", "Fixed Income", ("U.S. Inv Grade Corporate Bonds", "U.S. High Yield Bonds"), 0.00, 0.40, "Equal-weight U.S. Investment Grade Corporate Bonds and U.S. High Yield Bonds"),
    Asset("Income U.S. Securitized", "Fixed Income", ("U.S. Securitized",), 0.00, 0.03, "U.S. Securitized"),
    Asset("Fixed Income International", "Fixed Income", ("World ex-U.S. Government Bonds hedged", "Emerging Markets Sovereign Debt", "Emerging Markets Corporate Bonds"), 0.00, 0.10, "Equal-weight World ex-U.S. Government Bonds hedged, Emerging Markets Sovereign Debt, and Emerging Markets Corporate Bonds"),
    Asset("Other Fixed Income", "Fixed Income", ("U.S. Leveraged Loans", "Emerging Markets Local Currency Debt", "U.S. Muni High Yield"), 0.00, 0.01, "Equal-weight U.S. Leveraged Loans, Emerging Markets Local Currency Debt, and U.S. Muni High Yield"),
    Asset("Cash", "Cash", ("U.S. Cash",), CASH_TARGET, CASH_TARGET, "U.S. Cash"),
]


def target_volatility_range(profile: str) -> Tuple[float, float]:
    offset, half_width = TARGET_VOLATILITY_RULES[profile]
    midpoint = MODERATE_VOL_TARGET + offset
    return max(0.0, midpoint - half_width), max(0.0, midpoint + half_width)


def build_profiles() -> Dict[str, Dict[str, object]]:
    return {
        profile: {
            "target_volatility": target_volatility_range(profile),
            "category_bounds": category_bounds,
        }
        for profile, category_bounds in CATEGORY_BOUNDS.items()
    }


PROFILES: Dict[str, Dict[str, object]] = build_profiles()


def normalize_name(name: str) -> str:
    return " ".join(str(name).replace("\xa0", " ").split())


def jpm_matrix_path() -> Path:
    if JPM_MATRIX_FILE.exists():
        return JPM_MATRIX_FILE
    raise FileNotFoundError(f"Could not find JPM matrix file at {JPM_MATRIX_FILE}. Replace that file each quarter before refreshing.")


def load_jpm_source_data() -> Dict[str, object]:
    wb = load_workbook(jpm_matrix_path(), data_only=True)
    ws = wb["2026"] if "2026" in wb.sheetnames else wb.active

    row_by_name: Dict[str, int] = {}
    for row in range(9, ws.max_row + 1):
        value = ws.cell(row, 2).value
        if isinstance(value, str):
            row_by_name[normalize_name(value)] = row

    col_by_name: Dict[str, int] = {}
    for col in range(7, ws.max_column + 1):
        value = ws.cell(5, col).value
        if isinstance(value, str):
            col_by_name[normalize_name(value)] = col

    source_names = sorted(set(row_by_name).intersection(col_by_name))
    returns = {name: float(ws.cell(row_by_name[name], 6).value) / 100.0 for name in source_names}
    vols = {name: float(ws.cell(row_by_name[name], 5).value) / 100.0 for name in source_names}
    corr = np.eye(len(source_names))
    for i, row_name in enumerate(source_names):
        for j, col_name in enumerate(source_names):
            if i == j:
                corr[i, j] = 1.0
                continue
            value = ws.cell(row_by_name[row_name], col_by_name[col_name]).value
            if value is None:
                value = ws.cell(row_by_name[col_name], col_by_name[row_name]).value
            if value is None:
                raise ValueError(f"Missing JPM correlation for {row_name} / {col_name}")
            corr[i, j] = float(value)

    return {
        "source_names": source_names,
        "returns": returns,
        "vols": vols,
        "corr": corr,
        "source_index": {name: idx for idx, name in enumerate(source_names)},
    }


def asset_component_weights(asset: Asset) -> np.ndarray:
    return np.repeat(1.0 / len(asset.source_names), len(asset.source_names))


def asset_return(asset: Asset, jpm_data: Dict[str, object]) -> float:
    returns: Dict[str, float] = jpm_data["returns"]  # type: ignore[assignment]
    weights = asset_component_weights(asset)
    return float(sum(weights[i] * returns[normalize_name(name)] for i, name in enumerate(asset.source_names)))


def covariance_matrix(assets: List[Asset], jpm_data: Dict[str, object] | None = None) -> np.ndarray:
    data = jpm_data or load_jpm_source_data()
    source_index: Dict[str, int] = data["source_index"]  # type: ignore[assignment]
    vols_map: Dict[str, float] = data["vols"]  # type: ignore[assignment]
    source_corr: np.ndarray = data["corr"]  # type: ignore[assignment]
    source_names: List[str] = data["source_names"]  # type: ignore[assignment]
    source_vols = np.array([vols_map[name] for name in source_names])
    source_cov = np.outer(source_vols, source_vols) * source_corr

    exposure = np.zeros((len(assets), len(source_names)))
    for i, asset in enumerate(assets):
        weights = asset_component_weights(asset)
        for j, source_name in enumerate(asset.source_names):
            exposure[i, source_index[normalize_name(source_name)]] = weights[j]
    return exposure @ source_cov @ exposure.T


def asset_vol(asset: Asset, jpm_data: Dict[str, object]) -> float:
    cov = covariance_matrix([asset], jpm_data)
    return float(np.sqrt(cov[0, 0]))


def category_indices(assets: List[Asset]) -> Dict[str, List[int]]:
    out: Dict[str, List[int]] = {}
    for idx, asset in enumerate(assets):
        out.setdefault(asset.category, []).append(idx)
    return out


def random_feasible_asset_weights(profile_name: str, rng: np.random.Generator, draws: int) -> Iterable[np.ndarray]:
    cats = category_indices(ASSETS)
    min_w = np.array([a.min_weight for a in ASSETS])
    max_w = np.array([a.max_weight for a in ASSETS])
    for category_weights in random_feasible_category_weights(profile_name, rng, draws):
        category_weight_map = dict(zip(CATEGORIES, category_weights))
        for _attempt in range(400):
            weights = np.zeros(len(ASSETS))
            valid = True
            for category, indexes in cats.items():
                total = category_weight_map[category]
                idx = np.array(indexes)
                mins = min_w[idx]
                maxes = max_w[idx]
                if total < mins.sum() - 1e-9 or total > maxes.sum() + 1e-9:
                    valid = False
                    break
                if len(idx) == 1:
                    vals = np.array([total])
                else:
                    remaining_total = total - mins.sum()
                    vals = None
                    for _asset_attempt in range(400):
                        candidate = mins + rng.dirichlet(np.ones(len(idx))) * remaining_total
                        if np.all(candidate <= maxes + 1e-9):
                            vals = candidate
                            break
                    if vals is None:
                        valid = False
                        break
                weights[idx] = vals
            if valid:
                yield weights
                break


def optimize_anchor_portfolio(expected_returns: np.ndarray, cov: np.ndarray, profile_name: str = "Moderate") -> np.ndarray:
    profile = PROFILES[profile_name]
    bounds = [(asset.min_weight, asset.max_weight) for asset in ASSETS]
    cats = category_indices(ASSETS)
    vol_low, vol_high = profile["target_volatility"]  # type: ignore[index]
    constraints = [{"type": "eq", "fun": lambda w: np.sum(w) - 1.0}]
    for category, (low, high) in profile["category_bounds"].items():  # type: ignore[index]
        idx = cats[category]
        constraints.append({"type": "ineq", "fun": lambda w, idx=idx, low=low: np.sum(w[idx]) - low})
        constraints.append({"type": "ineq", "fun": lambda w, idx=idx, high=high: high - np.sum(w[idx])})
    constraints.append({"type": "ineq", "fun": lambda w, cov=cov, vol_low=vol_low: np.sqrt(float(w @ cov @ w)) - vol_low})
    constraints.append({"type": "ineq", "fun": lambda w, cov=cov, vol_high=vol_high: vol_high - np.sqrt(float(w @ cov @ w))})

    def objective(w: np.ndarray) -> float:
        return -float(w @ expected_returns)

    rng = np.random.default_rng(20260717)
    starts = list(random_feasible_asset_weights(profile_name, rng, 40))
    best_result = None
    for x0 in starts:
        result = minimize(objective, x0, method="SLSQP", bounds=bounds, constraints=constraints, options={"maxiter": 1500, "ftol": 1e-12})
        if result.success and (best_result is None or result.fun < best_result.fun):
            best_result = result
    if best_result is None:
        raise RuntimeError(f"Could not optimize anchor portfolio for {profile_name}")
    weights = np.where(best_result.x < 0.00005, 0.0, best_result.x)
    return weights / weights.sum()


def optimize_portfolio(
    expected_returns: np.ndarray,
    cov: np.ndarray,
    profile_name: str,
) -> np.ndarray:
    profile = PROFILES[profile_name]
    bounds = [(asset.min_weight, asset.max_weight) for asset in ASSETS]
    cats = category_indices(ASSETS)
    vol_low, vol_high = profile["target_volatility"]  # type: ignore[index]
    constraints = [{"type": "eq", "fun": lambda w: np.sum(w) - 1.0}]
    for category, (low, high) in profile["category_bounds"].items():  # type: ignore[index]
        idx = cats[category]
        constraints.append({"type": "ineq", "fun": lambda w, idx=idx, low=low: np.sum(w[idx]) - low})
        constraints.append({"type": "ineq", "fun": lambda w, idx=idx, high=high: high - np.sum(w[idx])})
    constraints.append({"type": "ineq", "fun": lambda w, cov=cov, vol_low=vol_low: np.sqrt(float(w @ cov @ w)) - vol_low})
    constraints.append({"type": "ineq", "fun": lambda w, cov=cov, vol_high=vol_high: vol_high - np.sqrt(float(w @ cov @ w))})

    def objective(w: np.ndarray) -> float:
        return -float(w @ expected_returns)

    rng = np.random.default_rng(20260717 + list(PROFILES).index(profile_name))
    starts = list(random_feasible_asset_weights(profile_name, rng, 60))

    best_result = None
    for x0 in starts:
        result = minimize(objective, x0, method="SLSQP", bounds=bounds, constraints=constraints, options={"maxiter": 1500, "ftol": 1e-12})
        if result.success and (best_result is None or result.fun < best_result.fun):
            best_result = result
    if best_result is None:
        no_vol_constraints = constraints[:-2]

        def target_gap(w: np.ndarray) -> float:
            vol = float(np.sqrt(w @ cov @ w))
            if vol < vol_low:
                return vol_low - vol
            if vol > vol_high:
                return vol - vol_high
            return 0.0

        def fallback_objective(w: np.ndarray) -> float:
            return target_gap(w) ** 2 - 0.0001 * float(w @ expected_returns)

        for x0 in starts:
            result = minimize(fallback_objective, x0, method="SLSQP", bounds=bounds, constraints=no_vol_constraints, options={"maxiter": 1500, "ftol": 1e-12})
            if result.success and (best_result is None or result.fun < best_result.fun):
                best_result = result
    if best_result is None:
        raise RuntimeError(f"Optimization failed for {profile_name}: no feasible allocation met the portfolio constraints")
    weights = best_result.x
    weights = np.where(weights < 0.00005, 0.0, weights)
    return weights / weights.sum()


def portfolio_stats(weights: np.ndarray, returns: np.ndarray, cov: np.ndarray) -> Tuple[float, float, float]:
    ret = float(weights @ returns)
    vol = float(np.sqrt(weights @ cov @ weights))
    sharpe = (ret - 0.031) / vol if vol else 0.0
    return ret, vol, sharpe


def target_vol_status(profile_name: str, vol: float) -> str:
    vol_low, vol_high = PROFILES[profile_name]["target_volatility"]  # type: ignore[index]
    if vol < vol_low - 0.0001:
        return "Below target"
    if vol > vol_high + 0.0001:
        return "Above target"
    return "In target"


def random_feasible_category_weights(profile_name: str, rng: np.random.Generator, draws: int) -> Iterable[np.ndarray]:
    profile = PROFILES[profile_name]
    cat_bounds = profile["category_bounds"]  # type: ignore[assignment]
    for _ in range(draws):
        for _attempt in range(200):
            cat_weights = {}
            remaining = 1.0
            feasible_draw = True
            fixed_total = 0.0
            variable_categories = []
            for cat in CATEGORIES:
                low, high = cat_bounds[cat]  # type: ignore[index]
                if abs(high - low) < 1e-12:
                    cat_weights[cat] = low
                    fixed_total += low
                else:
                    variable_categories.append(cat)
            remaining -= fixed_total
            for cat in variable_categories[:-1]:
                low, high = cat_bounds[cat]  # type: ignore[index]
                feasible_high = min(high, remaining)
                next_categories = variable_categories[variable_categories.index(cat) + 1:]
                feasible_low = max(low, remaining - sum(cat_bounds[c][1] for c in next_categories))  # type: ignore[index]
                if feasible_low > feasible_high:
                    feasible_draw = False
                    break
                cat_weights[cat] = rng.uniform(feasible_low, feasible_high)
                remaining -= cat_weights[cat]
            if not feasible_draw:
                continue
            if variable_categories:
                cat_weights[variable_categories[-1]] = remaining
            if any(cat not in cat_weights for cat in CATEGORIES):
                continue
            if all(cat_bounds[c][0] - 1e-9 <= cat_weights[c] <= cat_bounds[c][1] + 1e-9 for c in CATEGORIES):  # type: ignore[index]
                yield np.array([cat_weights[category] for category in CATEGORIES])
                break


def apply_percentile_target_volatility(returns: np.ndarray, cov: np.ndarray) -> None:
    for profile_name in PROFILES:
        rng = np.random.default_rng(8901 + list(PROFILES).index(profile_name))
        vols = []
        for weights in random_feasible_asset_weights(profile_name, rng, 2500):
            vols.append(float(np.sqrt(weights @ cov @ weights)))
        if not vols:
            continue
        midpoint = float(np.percentile(vols, VOLATILITY_PERCENTILES[profile_name]))
        half_width = TARGET_VOLATILITY_RULES[profile_name][1]
        PROFILES[profile_name]["target_volatility"] = (max(0.0, midpoint - half_width), max(0.0, midpoint + half_width))


def pct(x: float) -> float:
    return round(x, 6)


def calculate_results() -> Dict[str, object]:
    jpm_data = load_jpm_source_data()
    returns = np.array([asset_return(a, jpm_data) for a in ASSETS])
    cov = covariance_matrix(ASSETS, jpm_data)
    apply_percentile_target_volatility(returns, cov)
    mvo = {}
    for profile in PROFILES:
        weights = optimize_portfolio(returns, cov, profile)
        mvo[profile] = {"weights": weights, "stats": portfolio_stats(weights, returns, cov)}

    scenarios = {}
    for scenario in ["Bad", "Median", "Good"]:
        vols = np.sqrt(np.diag(cov))
        if scenario == "Bad":
            s_returns = returns - 0.50 * vols
        elif scenario == "Good":
            s_returns = returns + 0.50 * vols
        else:
            s_returns = returns
        scenarios[scenario] = {}
        for profile in PROFILES:
            weights = optimize_portfolio(s_returns, cov, profile)
            scenarios[scenario][profile] = {"weights": weights, "stats": portfolio_stats(weights, s_returns, cov)}

    return {
        "mvo": mvo,
        "scenarios": scenarios,
        "returns": returns,
        "cov": cov,
        "jpm_data": jpm_data,
    }


def style_sheet(ws, freeze: str | None = None) -> None:
    ws.sheet_view.showGridLines = False
    if freeze:
        ws.freeze_panes = freeze
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=Side(style="hair", color="D9E2EC"))
    for col in range(1, ws.max_column + 1):
        max_len = 0
        for cell in ws[get_column_letter(col)]:
            max_len = max(max_len, len(str(cell.value)) if cell.value is not None else 0)
        ws.column_dimensions[get_column_letter(col)].width = min(max(max_len + 2, 10), 34)


def header(ws, row: int, start_col: int, end_col: int, title: str) -> None:
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    cell = ws.cell(row=row, column=start_col, value=title)
    cell.fill = PatternFill("solid", fgColor="17324D")
    cell.font = Font(color="FFFFFF", bold=True)
    cell.alignment = Alignment(horizontal="left")


def write_weights_table(ws, start_row: int, title: str, method_data: Dict[str, Dict[str, object]]) -> int:
    header(ws, start_row, 1, 11, title)
    row = start_row + 1
    ws.cell(row=row, column=1, value="Portfolio")
    ws.cell(row=row, column=2, value="Expected Return")
    ws.cell(row=row, column=3, value="Volatility")
    ws.cell(row=row, column=4, value="Sharpe vs Cash")
    ws.cell(row=row, column=5, value="Equity")
    ws.cell(row=row, column=6, value="Fixed Income")
    ws.cell(row=row, column=7, value="Alternatives")
    ws.cell(row=row, column=8, value="Cash")
    ws.cell(row=row, column=9, value="Target Vol Min")
    ws.cell(row=row, column=10, value="Target Vol Max")
    ws.cell(row=row, column=11, value="Target Status")
    for cell in ws[row]:
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.font = Font(bold=True)
    row += 1
    cats = category_indices(ASSETS)
    for profile, data in method_data.items():
        weights = data["weights"]  # type: ignore[assignment]
        if "stats" in data:
            stats = data["stats"]  # type: ignore[assignment]
        else:
            stats = (data["return"], data["vol"], data["sharpe"])  # type: ignore[index]
        ws.cell(row=row, column=1, value=profile)
        ws.cell(row=row, column=2, value=pct(stats[0]))
        ws.cell(row=row, column=3, value=pct(stats[1]))
        ws.cell(row=row, column=4, value=pct(stats[2]))
        for offset, cat in enumerate(["Equity", "Fixed Income", "Alternatives", "Cash"], start=5):
            ws.cell(row=row, column=offset, value=pct(float(np.sum(weights[cats[cat]]))))
        target_low, target_high = PROFILES[profile]["target_volatility"]  # type: ignore[index]
        ws.cell(row=row, column=9, value=target_low)
        ws.cell(row=row, column=10, value=target_high)
        ws.cell(row=row, column=11, value=target_vol_status(profile, stats[1]))
        row += 1
    for r in range(start_row + 2, row):
        for c in list(range(2, 4)) + list(range(5, 11)):
            ws.cell(r, c).number_format = "0.0%"
        ws.cell(r, 4).number_format = "0.00"
    return row + 2


def build_workbook(results: Dict[str, object]) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    input_font = Font(color="0000FF")
    formula_font = Font(color="000000")
    link_font = Font(color="008000")

    ws["A1"] = "Paul Asset Allocation Optimization Model"
    ws["A1"].font = Font(size=16, bold=True, color="17324D")
    ws["A3"] = "Purpose"
    ws["B3"] = "Optimize five risk-level portfolios using compound return, volatility, and correlation assumptions with visible constraints."
    ws["A4"] = "Methods"
    ws["B4"] = "Main allocation is the target-volatility MVO result. Monte Carlo simulation is currently removed from all calculations."
    ws["A5"] = "Important"
    ws["B5"] = "This is a planning model, not investment advice. Correlations and constraints are starter assumptions that should be reviewed."
    ws["A7"] = "Model Status"
    ws["B7"] = "=IF(Checks!G2=\"OK\",\"OK\",\"Review checks\")"
    ws["B7"].font = formula_font

    row = 9
    row = write_weights_table(ws, row, "Main Recommended Allocation: MVO", results["mvo"])  # type: ignore[arg-type]

    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = "Main Recommended Overall Allocation"
    chart.y_axis.title = "Portfolio"
    chart.x_axis.title = "Weight"
    data = Reference(ws, min_col=5, max_col=8, min_row=10, max_row=15)
    cats = Reference(ws, min_col=1, min_row=11, max_row=15)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 8
    chart.width = 18
    ws.add_chart(chart, "J9")

    row = write_weights_table(ws, row, "Target Volatility Optimized Overall Allocation", results["mvo"])  # type: ignore[arg-type]

    ws2 = wb.create_sheet("Asset Inputs")
    headers = [
        "Asset",
        "Category",
        "Source Name(s)",
        "Compound Return",
        "Volatility",
        "Min Weight",
        "Max Weight",
        "Source Mapping",
    ]
    ws2.append(headers)
    jpm_data = results["jpm_data"]  # type: ignore[assignment]
    for asset in ASSETS:
        ws2.append([
            asset.name,
            asset.category,
            ", ".join(asset.source_names),
            asset_return(asset, jpm_data),
            asset_vol(asset, jpm_data),
            asset.min_weight,
            asset.max_weight,
            asset.source_note,
        ])
    for row_cells in ws2.iter_rows(min_row=2, max_col=8):
        for cell in row_cells[3:7]:
            cell.number_format = "0.0%"
            cell.font = input_font

    ws3 = wb.create_sheet("MVO Detail")
    ws3.append(["Asset", "Category"] + list(PROFILES))
    for i, asset in enumerate(ASSETS):
        ws3.append([asset.name, asset.category] + [pct(results["mvo"][p]["weights"][i]) for p in PROFILES])  # type: ignore[index]
    ws3.append(["Total", "Check"] + [f"=SUM({get_column_letter(c)}2:{get_column_letter(c)}{len(ASSETS) + 1})" for c in range(3, 3 + len(PROFILES))])
    for r in range(2, ws3.max_row + 1):
        for c in range(3, ws3.max_column + 1):
            ws3.cell(r, c).number_format = "0.0%"

    ws5 = wb.create_sheet("Scenario Detail")
    scenario_row = 1
    for scenario in ["Bad", "Median", "Good"]:
        header(ws5, scenario_row, 1, 8, f"{scenario} Economy Optimized Overall Allocation")
        scenario_row += 1
        ws5.append(["Portfolio", "Expected Return", "Volatility", "Sharpe", "Equity", "Fixed Income", "Alternatives", "Cash", "Target Vol Min", "Target Vol Max", "Target Status"])
        cats = category_indices(ASSETS)
        for profile in PROFILES:
            data = results["scenarios"][scenario][profile]  # type: ignore[index]
            weights = data["weights"]
            stats = data["stats"]
            ws5.append([
                profile,
                pct(stats[0]),
                pct(stats[1]),
                pct(stats[2]),
                pct(float(np.sum(weights[cats["Equity"]]))),
                pct(float(np.sum(weights[cats["Fixed Income"]]))),
                pct(float(np.sum(weights[cats["Alternatives"]]))),
                pct(float(np.sum(weights[cats["Cash"]]))),
                PROFILES[profile]["target_volatility"][0],  # type: ignore[index]
                PROFILES[profile]["target_volatility"][1],  # type: ignore[index]
                target_vol_status(profile, stats[1]),
            ])
        scenario_row = ws5.max_row + 3
    for r in range(1, ws5.max_row + 1):
        for c in range(2, 11):
            if isinstance(ws5.cell(r, c).value, (int, float)):
                ws5.cell(r, c).number_format = "0.0%"
        if ws5.cell(r, 4).value and isinstance(ws5.cell(r, 4).value, float):
            ws5.cell(r, 4).number_format = "0.00"

    ws6 = wb.create_sheet("Constraints")
    ws6.append(["Portfolio", "Category", "Minimum", "Maximum", "Target Vol Min", "Target Vol Max", "Notes"])
    for profile, config in PROFILES.items():
        for cat, (low, high) in config["category_bounds"].items():  # type: ignore[index]
            target_low, target_high = config["target_volatility"]  # type: ignore[index]
            ws6.append([profile, cat, low, high, target_low, target_high, "Editable policy band used by optimizer script"])
    ws6.append([])
    ws6.append(["Asset", "Sleeve Range Min", "Sleeve Range Max", "Notes"])
    for asset in ASSETS:
        ws6.append([asset.name, asset.min_weight, asset.max_weight, "Direct asset-level bounds used by optimizer"])
    for r in range(2, ws6.max_row + 1):
        for c in [2, 3, 4, 5, 6]:
            ws6.cell(r, c).number_format = "0.0%"
    for cell in ws6["B"][1:] + ws6["C"][1:] + ws6["D"][1:] + ws6["E"][1:] + ws6["F"][1:]:
        cell.font = input_font

    ws7 = wb.create_sheet("Checks")
    ws7.append(["Check", "Actual", "Expected", "Difference", "Tolerance", "Status", "Model Status"])
    row = 2
    for profile_idx, profile in enumerate(PROFILES, start=3):
        col = get_column_letter(profile_idx)
        ws7.append([
            f"{profile} MVO weights sum to 100%",
            f"='MVO Detail'!{col}17",
            1.0,
            f"=B{row}-C{row}",
            0.0001,
            f"=IF(ABS(D{row})<=E{row},\"OK\",\"Review\")",
            f"=IF(COUNTIF(F2:F{row},\"Review\")=0,\"OK\",\"Review\")",
        ])
        row += 1
    cash_asset_row = 1 + next(i for i, asset in enumerate(ASSETS, start=1) if asset.category == "Cash")
    for profile_idx, profile in enumerate(PROFILES, start=3):
        col = get_column_letter(profile_idx)
        ws7.append([
            f"{profile} MVO cash equals 1%",
            f"='MVO Detail'!{col}{cash_asset_row}",
            CASH_TARGET,
            f"=B{row}-C{row}",
            0.0001,
            f"=IF(ABS(D{row})<=E{row},\"OK\",\"Review\")",
            "",
        ])
        row += 1
    ws7.append(["All MVO sub-weights are non-negative", "=MIN('MVO Detail'!C2:G16)", 0.0, f"=B{row}-C{row}", 0.0, f"=IF(B{row}>=0,\"OK\",\"Review\")", ""])
    for r in range(2, ws7.max_row + 1):
        for c in [2, 3, 4, 5]:
            ws7.cell(r, c).number_format = "0.000%"

    ws8 = wb.create_sheet("Sources and Setup")
    ws8.append(["Item", "Current Treatment", "What to Decide Next"])
    setup_rows = [
        ("Source file", f"Matrix file read from {jpm_matrix_path()}; compound return, annualized volatility, and correlations are used.", "Confirm source file version each quarter."),
        ("PDF attachments", "Scan_260717_102740.pdf and Scan_260717_102947.pdf are present but no longer used in the optimizer.", "Keep or remove from source archive."),
        ("Return source", "Compound return only.", "Confirm this remains the preferred expected return basis."),
        ("Risk model", "Uses annualized volatility and correlation matrix from matrix-usd.xlsx.", "Review asset mappings for grouped fixed income sleeves."),
        ("Monte Carlo", "Removed from all calculations for now.", "Redesign simulation methodology before adding it back."),
        ("Five portfolios", "Category bands and target volatility ranges are explicit on Constraints tab; cash is hard-coded with both minimum and maximum at 1% in every model.", "Approve or revise the equity/fixed income/alternatives policy bands and target volatility bands."),
        ("Tax/location", "Not modeled.", "Decide taxable vs qualified account treatment, muni bond role, and tax-aware placement."),
        ("Liquidity/implementation", "No ETF/mutual fund tickers selected yet.", "Map each style sleeve to actual funds and expense ratios."),
        ("Rebalancing", "Not modeled.", "Choose annual, semiannual, or tolerance-band rebalancing."),
    ]
    for row_data in setup_rows:
        ws8.append(row_data)

    for sheet in wb.worksheets:
        if sheet.max_row >= 1:
            for cell in sheet[1]:
                cell.fill = PatternFill("solid", fgColor="17324D")
                cell.font = Font(color="FFFFFF", bold=True)
        style_sheet(sheet, "A2" if sheet.max_row > 8 else None)

    for sheet in ["Summary", "Asset Inputs", "MVO Detail", "Scenario Detail", "Constraints"]:
        wsx = wb[sheet]
        for row_cells in wsx.iter_rows():
            for cell in row_cells:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    cell.font = formula_font
                elif sheet in {"Asset Inputs", "Constraints"} and cell.row > 1:
                    if cell.column in {4, 5, 6, 7}:
                        cell.font = input_font
    ws8["A2"].font = link_font

    wb.save(OUTPUT_FILE)


def main() -> None:
    results = calculate_results()
    build_workbook(results)
    print(OUTPUT_FILE)


if __name__ == "__main__":
    main()
