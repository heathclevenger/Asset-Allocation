from __future__ import annotations

import csv
import json
from pathlib import Path

import numpy as np
from openpyxl import load_workbook

import portfolio_optimizer as optimizer


ROOT = Path(__file__).resolve().parents[1]
WEB_DATA = ROOT / "web" / "data" / "model-data.json"
BLACKROCK_ASSUMPTIONS_FILE = ROOT / "inputs" / "blackrock-capital-market-assumptions.xlsx"
VANGUARD_ASSUMPTIONS_FILE = ROOT / "inputs" / "vanguard_2026_assumptions.csv"
INVESCO_ASSUMPTIONS_FILE = ROOT / "inputs" / "invesco_2026_assumptions.csv"
MSCI_ASSUMPTIONS_FILE = ROOT / "inputs" / "msci_2026_assumptions.csv"
CAPITAL_GROUP_FILE = ROOT / "inputs" / "capital_group_2026.xlsx"
ASSET_ALLOCATION_INTERACTIVE_FILE = ROOT / "inputs" / "Asset-Allocation-Interactive-Data.xlsx"


def parse_assumption_percent(value: str) -> float:
    text = str(value).strip()
    if text.endswith("%"):
        return float(text[:-1]) / 100.0
    number = float(text)
    return number / 100.0 if abs(number) > 1 else number


def load_vanguard_assumptions() -> dict[str, dict[str, float | str]]:
    if not VANGUARD_ASSUMPTIONS_FILE.exists():
        return {}
    out: dict[str, dict[str, float | str]] = {}
    with VANGUARD_ASSUMPTIONS_FILE.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            model_asset = optimizer.normalize_name(row.get("Model Asset", ""))
            provider_asset = row.get("Provider Asset", "").strip()
            if not model_asset or not provider_asset:
                continue
            out[model_asset] = {
                "return": parse_assumption_percent(row.get("Return", "0")),
                "volatility": parse_assumption_percent(row.get("Volatility", "0")),
                "sourceMapping": provider_asset,
            }
    return out


JPM_FALLBACK_TO_AVERAGE = {"U.S. Growth"}
CAPITAL_GROUP_FALLBACK_TO_AVERAGE = {
    "U.S. Value",
    "U.S. Growth",
    "U.S. Income",
    "U.S. Quality",
    "U.S. REITs",
    "Commodities",
}
ASSET_ALLOCATION_INTERACTIVE_FALLBACK_TO_AVERAGE = {
    "U.S. Growth",
    "U.S. Quality",
}
INVESCO_FALLBACK_TO_AVERAGE = {
    "U.S. Mid Cap",
    "U.S. Value",
    "U.S. Growth",
    "U.S. Income",
    "U.S. Quality",
    "U.S. REITs",
    "Cash",
}
FALLBACK_RETURN_OVERRIDES = {
    "U.S. Growth": 0.0700,
    "U.S. Income": 0.0600,
    "U.S. Quality": 0.0610,
}


def load_msci_assumptions() -> dict[str, dict[str, float | str]]:
    if not MSCI_ASSUMPTIONS_FILE.exists():
        return {}
    out: dict[str, dict[str, float | str]] = {}
    with MSCI_ASSUMPTIONS_FILE.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            model_asset = optimizer.normalize_name(row.get("Model Asset", ""))
            provider_asset = row.get("Provider Asset", "").strip()
            if not model_asset or not provider_asset:
                continue
            out[model_asset] = {
                "return": parse_assumption_percent(row.get("Return", "0")),
                "volatility": parse_assumption_percent(row.get("Volatility", "0")),
                "sourceMapping": provider_asset,
            }
    return out


def load_capital_group_assumptions() -> dict[str, dict[str, float | str]]:
    if not CAPITAL_GROUP_FILE.exists():
        return {}
    out: dict[str, dict[str, float | str]] = {}
    wb = load_workbook(CAPITAL_GROUP_FILE, data_only=True, read_only=True)
    ws = wb["Assumptions"]
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
    index = {header: position for position, header in enumerate(headers)}
    for row in ws.iter_rows(min_row=2, values_only=True):
        model_asset = optimizer.normalize_name(row[index["Model Asset"]] or "")
        provider_asset = str(row[index["Provider Asset"]] or "").strip()
        expected_return = row[index["Return"]]
        volatility = row[index["Volatility"]]
        if not model_asset or not provider_asset or expected_return in (None, "") or volatility in (None, ""):
            continue
        out[model_asset] = {
            "return": parse_assumption_percent(str(expected_return)),
            "volatility": parse_assumption_percent(str(volatility)),
            "sourceMapping": provider_asset,
        }
    return out


def load_capital_group_correlation() -> dict[str, dict[str, float]]:
    if not CAPITAL_GROUP_FILE.exists():
        return {}
    wb = load_workbook(CAPITAL_GROUP_FILE, data_only=True, read_only=True)
    ws = wb["Correlations"]
    headers = [optimizer.normalize_name(cell.value) for cell in ws[1][1:]]
    out: dict[str, dict[str, float]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_name = optimizer.normalize_name(row[0])
        if not row_name:
            continue
        out[row_name] = {
            header: float(value)
            for header, value in zip(headers, row[1:])
            if header and value not in (None, "")
        }
    return out


def load_asset_allocation_interactive_assumptions() -> dict[str, dict[str, float | str]]:
    if not ASSET_ALLOCATION_INTERACTIVE_FILE.exists():
        return {}
    wb = load_workbook(ASSET_ALLOCATION_INTERACTIVE_FILE, data_only=True, read_only=True)
    ws = wb["Model Mapping"]
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
    index = {header: position for position, header in enumerate(headers)}
    out: dict[str, dict[str, float | str]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        model_asset = optimizer.normalize_name(row[index["Model Asset"]] or "")
        provider_asset = str(row[index["Provider Asset"]] or "").strip()
        expected_return = row[index["Return"]]
        volatility = row[index["Volatility"]]
        if not model_asset or not provider_asset or expected_return in (None, "") or volatility in (None, ""):
            continue
        out[model_asset] = {
            "return": parse_assumption_percent(str(expected_return)),
            "volatility": parse_assumption_percent(str(volatility)),
            "sourceMapping": provider_asset,
        }
    return out


def load_asset_allocation_interactive_correlation() -> dict[str, dict[str, float]]:
    if not ASSET_ALLOCATION_INTERACTIVE_FILE.exists():
        return {}
    wb = load_workbook(ASSET_ALLOCATION_INTERACTIVE_FILE, data_only=True, read_only=True)
    ws = wb["Expected Correlations"]
    headers = [optimizer.normalize_name(cell.value) for cell in ws[4][2:] if cell.value not in (None, "")]
    out: dict[str, dict[str, float]] = {}
    for row in ws.iter_rows(min_row=5, values_only=True):
        row_name = optimizer.normalize_name(row[1])
        if not row_name or row_name in out:
            continue
        values = row[2:2 + len(headers)]
        if not values or all(value in (None, "") for value in values):
            continue
        out[row_name] = {
            header: float(value)
            for header, value in zip(headers, values)
            if header and value not in (None, "")
        }
        if len(out) >= len(headers):
            break
    return out
BLACKROCK_FALLBACK_TO_AVERAGE = {
    "U.S. Value",
    "U.S. Growth",
    "U.S. Income",
    "U.S. Quality",
    "Commodities",
    "US Short Treasuries",
}

CAPITAL_GROUP_CORRELATION_MAP = {
    "U.S. Large Cap": "U.S. equity",
    "U.S. Mid Cap": ("U.S. equity", "U.S. small-cap equity"),
    "U.S. Small Cap": "U.S. small-cap equity",
    "International Developed Equity": "Non-U.S. developed markets equity",
    "Emerging Markets Equity": "Emerging markets equity",
    "US Short Treasuries": "U.S. Treasury short term",
    "US Intermediate Treasuries": "U.S. Treasury intermediate term",
    "US Long Treasuries": "U.S. Treasury long term",
    "Investment Grade Corporate": "U.S. corporate",
    "High Yield": "U.S. high yield",
    "International Fixed Income (H)": "Non-U.S. global aggregate",
    "Cash": "Cash (USD)",
}

ASSET_ALLOCATION_INTERACTIVE_CORRELATION_MAP = {
    "U.S. Large Cap": "US Large",
    "U.S. Mid Cap": ("US Large", "US Small"),
    "U.S. Small Cap": "US Small",
    "U.S. Value": "US Large Value",
    "U.S. Growth": "US Large Growth",
    "U.S. Income": "US Large RAFI",
    "International Developed Equity": "Dev ex US Large",
    "Emerging Markets Equity": "Emerging Markets",
    "U.S. REITs": "REITs",
    "Commodities": "Commodities",
    "US Short Treasuries": "US Treasury Short",
    "US Intermediate Treasuries": "US Treasury Intermediate",
    "US Long Treasuries": "US Treasury Long",
    "Investment Grade Corporate": "US Corporate Intermediate",
    "High Yield": "US High Yield",
    "International Fixed Income (H)": "Global ex US Aggregate (Hedged)",
    "Cash": "United States Cash",
}


def load_invesco_assumptions() -> dict[str, dict[str, float | str]]:
    if not INVESCO_ASSUMPTIONS_FILE.exists():
        return {}
    out: dict[str, dict[str, float | str]] = {}
    with INVESCO_ASSUMPTIONS_FILE.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            model_asset = optimizer.normalize_name(row.get("Model Asset", ""))
            provider_asset = row.get("Provider Asset", "").strip()
            expected_return = row.get("Arithmetic Return", "").strip()
            expected_risk = row.get("Expected Risk", "").strip()
            if not model_asset or not provider_asset or not expected_return or not expected_risk:
                continue
            out[model_asset] = {
                "return": parse_assumption_percent(expected_return),
                "volatility": parse_assumption_percent(expected_risk),
                "sourceMapping": provider_asset,
            }
    return out


def load_blackrock_source_data() -> dict[str, dict[str, float]]:
    if not BLACKROCK_ASSUMPTIONS_FILE.exists():
        return {}
    wb = load_workbook(BLACKROCK_ASSUMPTIONS_FILE, data_only=True, read_only=True)
    ws = wb["Starting point "] if "Starting point " in wb.sheetnames else wb.active
    out: dict[str, dict[str, float]] = {}
    for asset_class, asset_name, expected_return, volatility in ws.iter_rows(min_row=4, max_col=4, values_only=True):
        if not isinstance(asset_name, str):
            continue
        if not isinstance(expected_return, (int, float)) or not isinstance(volatility, (int, float)):
            continue
        out[optimizer.normalize_name(asset_name)] = {
            "return": float(expected_return),
            "volatility": float(volatility),
        }
    return out


def blackrock_assumption(source: dict[str, dict[str, float]], name: str, source_mapping: str) -> dict[str, float | str] | None:
    row = source.get(optimizer.normalize_name(name))
    if row is None:
        return None
    return {
        "return": row["return"],
        "volatility": row["volatility"],
        "sourceMapping": source_mapping,
    }


def blackrock_average_assumption(source: dict[str, dict[str, float]], names: list[str], source_mapping: str) -> dict[str, float | str] | None:
    rows = [source.get(optimizer.normalize_name(name)) for name in names]
    if any(row is None for row in rows):
        return None
    valid_rows = [row for row in rows if row is not None]
    return {
        "return": sum(row["return"] for row in valid_rows) / len(valid_rows),
        "volatility": sum(row["volatility"] for row in valid_rows) / len(valid_rows),
        "sourceMapping": source_mapping,
    }


def build_blackrock_assumptions(jpm_assets: list[dict]) -> list[dict | None]:
    source = load_blackrock_source_data()
    mappings = {
        "U.S. Large Cap": lambda: blackrock_assumption(source, "US large cap equities", "BlackRock US large cap equities"),
        "U.S. Mid Cap": lambda: blackrock_average_assumption(
            source,
            ["US large cap equities", "US small cap equities"],
            "BlackRock average of US large cap equities and US small cap equities",
        ),
        "U.S. Small Cap": lambda: blackrock_assumption(source, "US small cap equities", "BlackRock US small cap equities"),
        "International Developed Equity": lambda: blackrock_assumption(source, "Global ex-US large cap equities", "BlackRock Global ex-US large cap equities"),
        "Emerging Markets Equity": lambda: blackrock_assumption(source, "Emerging large cap equities", "BlackRock Emerging large cap equities"),
        "U.S. REITs": lambda: blackrock_assumption(source, "Listed REITs", "BlackRock Listed REITs"),
        "US Intermediate Treasuries": lambda: blackrock_assumption(source, "US government (10- years)", "BlackRock US government (10- years)"),
        "US Long Treasuries": lambda: blackrock_assumption(source, "US government (10+ years)", "BlackRock US government (10+ years)"),
        "Investment Grade Corporate": lambda: blackrock_assumption(source, "US credit (all maturities)", "BlackRock US credit (all maturities)"),
        "High Yield": lambda: blackrock_assumption(source, "US high yield", "BlackRock US high yield"),
        "International Fixed Income (H)": lambda: blackrock_assumption(source, "Global ex-US treasuries (hedged)", "BlackRock Global ex-US treasuries (hedged)"),
        "Cash": lambda: blackrock_assumption(source, "US cash", "BlackRock US cash"),
    }

    out = []
    for asset in jpm_assets:
        assumption = mappings.get(asset["name"], lambda: None)()
        if assumption is None:
            out.append(None)
            continue
        out.append({
            **asset,
            "return": assumption["return"],
            "volatility": assumption["volatility"],
            "sourceMapping": assumption["sourceMapping"],
        })
    return out


def average_assumption_sets(source_sets: list[list[dict | None]], fallback_by_source: list[set[str]]) -> list[dict]:
    first_pass = []
    base_assets = source_sets[0]
    for index, base_asset in enumerate(base_assets):
        if base_asset is None:
            continue
        asset = dict(base_asset)
        available = []
        for source_index, source_assets in enumerate(source_sets):
            source_asset = source_assets[index]
            if source_asset is not None and asset["name"] not in fallback_by_source[source_index]:
                available.append(source_asset)
        if not available:
            available.append(base_asset)
        asset["return"] = sum(source["return"] for source in available) / len(available)
        asset["volatility"] = sum(source["volatility"] for source in available) / len(available)
        asset["_distinctSourceCount"] = len(available)
        first_pass.append(asset)

    averaged = []
    for index, asset in enumerate(first_pass):
        fallback_return = FALLBACK_RETURN_OVERRIDES.get(asset["name"])
        if fallback_return is not None:
            return_values = []
            for source_index, source_assets in enumerate(source_sets):
                source_asset = source_assets[index]
                if source_asset is not None and asset["name"] not in fallback_by_source[source_index]:
                    return_values.append(source_asset["return"])
                elif asset["name"] in fallback_by_source[source_index]:
                    return_values.append(fallback_return)
            if return_values:
                asset["return"] = sum(return_values) / len(return_values)
        asset["sourceMapping"] = "Average of available source assumptions"
        averaged.append(asset)
    return averaged


def apply_average_fallback(source_assets: list[dict | None], average_assets: list[dict], source_name: str, fallback_names: set[str] | None = None) -> list[dict]:
    fallback_names = fallback_names or set()
    out = []
    for source_asset, average_asset in zip(source_assets, average_assets):
        if source_asset is not None and source_asset["name"] not in fallback_names:
            out.append(source_asset)
            continue
        asset = dict(average_asset)
        fallback_return = FALLBACK_RETURN_OVERRIDES.get(asset["name"])
        if fallback_return is not None:
            asset["return"] = fallback_return
            asset["sourceMapping"] = f"{source_name} was not given for this sleeve; fallback return input and AVERAGE volatility assumptions are used."
        else:
            asset["sourceMapping"] = f"{source_name} was not given for this sleeve; AVERAGE assumptions are used."
        out.append(asset)
    return out


def mapped_correlation_names(mapping_by_asset: dict[str, object], asset_name: str) -> list[str]:
    mapping = mapping_by_asset.get(asset_name)
    if mapping is None:
        return []
    if isinstance(mapping, tuple):
        return [optimizer.normalize_name(name) for name in mapping]
    return [optimizer.normalize_name(mapping)]


def source_correlation(source_corr: dict[str, dict[str, float]], left_names: list[str], right_names: list[str]) -> float | None:
    values = []
    for left_name in left_names:
        for right_name in right_names:
            value = source_corr.get(left_name, {}).get(right_name)
            if value is None:
                value = source_corr.get(right_name, {}).get(left_name)
            if value is not None:
                values.append(float(value))
    if not values:
        return None
    return sum(values) / len(values)


def blend_available_correlations(jpm_corr: np.ndarray, assets: list[dict]) -> np.ndarray:
    providers = [
        (load_capital_group_correlation(), CAPITAL_GROUP_CORRELATION_MAP),
        (load_asset_allocation_interactive_correlation(), ASSET_ALLOCATION_INTERACTIVE_CORRELATION_MAP),
    ]
    blended = np.array(jpm_corr, copy=True)
    for i, left_asset in enumerate(assets):
        for j, right_asset in enumerate(assets):
            if i == j:
                blended[i, j] = 1.0
                continue
            values = [float(jpm_corr[i, j])]
            for provider_corr, provider_map in providers:
                left_names = mapped_correlation_names(provider_map, left_asset["name"])
                right_names = mapped_correlation_names(provider_map, right_asset["name"])
                if not left_names or not right_names:
                    continue
                provider_value = source_correlation(provider_corr, left_names, right_names)
                if provider_value is not None:
                    values.append(provider_value)
            blended[i, j] = sum(values) / len(values)
    return (blended + blended.T) / 2.0


def blend_capital_group_correlation(jpm_corr: np.ndarray, assets: list[dict]) -> np.ndarray:
    capital_corr = load_capital_group_correlation()
    if not capital_corr:
        return jpm_corr
    def mapped_names(asset_name: str) -> list[str]:
        mapping = CAPITAL_GROUP_CORRELATION_MAP.get(asset_name)
        if mapping is None:
            return []
        if isinstance(mapping, tuple):
            return [optimizer.normalize_name(name) for name in mapping]
        return [optimizer.normalize_name(mapping)]

    def capital_correlation(left_names: list[str], right_names: list[str]) -> float | None:
        values = []
        for left_name in left_names:
            for right_name in right_names:
                value = capital_corr.get(left_name, {}).get(right_name)
                if value is None:
                    value = capital_corr.get(right_name, {}).get(left_name)
                if value is not None:
                    values.append(float(value))
        if not values:
            return None
        return sum(values) / len(values)

    blended = np.array(jpm_corr, copy=True)
    for i, left_asset in enumerate(assets):
        left_names = mapped_names(left_asset["name"])
        if not left_names:
            continue
        for j, right_asset in enumerate(assets):
            if i == j:
                blended[i, j] = 1.0
                continue
            right_names = mapped_names(right_asset["name"])
            if not right_names:
                continue
            capital_value = capital_correlation(left_names, right_names)
            if capital_value is None:
                continue
            blended[i, j] = (float(jpm_corr[i, j]) + float(capital_value)) / 2.0
    return (blended + blended.T) / 2.0


def nearest_valid_correlation(corr: np.ndarray) -> np.ndarray:
    cleaned = (corr + corr.T) / 2.0
    values, vectors = np.linalg.eigh(cleaned)
    values = np.maximum(values, 1e-8)
    cleaned = (vectors @ np.diag(values) @ vectors.T)
    diagonal = np.sqrt(np.maximum(np.diag(cleaned), 1e-8))
    cleaned = cleaned / np.outer(diagonal, diagonal)
    cleaned = np.clip((cleaned + cleaned.T) / 2.0, -1.0, 1.0)
    np.fill_diagonal(cleaned, 1.0)
    return cleaned


def main() -> None:
    jpm_data = optimizer.load_jpm_source_data()
    returns = np.array([optimizer.asset_return(asset, jpm_data) for asset in optimizer.ASSETS])
    cov = optimizer.covariance_matrix(optimizer.ASSETS, jpm_data)
    vols = np.sqrt(np.diag(cov))
    corr = cov / np.outer(vols, vols)
    corr = np.nan_to_num(corr, nan=0.0)
    volatility_model = {
        "mode": "feasiblePercentile",
        "halfWidth": max(half_width for _offset, half_width in optimizer.TARGET_VOLATILITY_RULES.values()),
        "profiles": {},
    }
    for profile, (offset, half_width) in optimizer.TARGET_VOLATILITY_RULES.items():
        volatility_model["profiles"][profile] = {
            "percentile": optimizer.VOLATILITY_PERCENTILES[profile],
            "halfWidth": volatility_model["halfWidth"],
        }

    jpm_assets = [
        {
            "name": asset.name,
            "category": asset.category,
            "sourceNames": list(asset.source_names),
            "return": float(returns[i]),
            "volatility": float(vols[i]),
            "minWeight": asset.min_weight,
            "maxWeight": asset.max_weight,
            "sourceMapping": asset.source_note,
        }
        for i, asset in enumerate(optimizer.ASSETS)
    ]
    vanguard_assumptions = load_vanguard_assumptions()
    raw_vanguard_assets = []
    for asset in jpm_assets:
        normalized_asset_name = optimizer.normalize_name(asset["name"])
        if normalized_asset_name in vanguard_assumptions:
            assumption = vanguard_assumptions[normalized_asset_name]
            raw_vanguard_assets.append({
                **asset,
                "return": assumption["return"],
                "volatility": assumption["volatility"],
                "sourceMapping": assumption["sourceMapping"],
            })
        else:
            raw_vanguard_assets.append(None)
    raw_blackrock_assets = build_blackrock_assumptions(jpm_assets)
    capital_group_assumptions = load_capital_group_assumptions()
    raw_capital_group_assets = []
    for asset in jpm_assets:
        normalized_asset_name = optimizer.normalize_name(asset["name"])
        if normalized_asset_name in capital_group_assumptions:
            assumption = capital_group_assumptions[normalized_asset_name]
            raw_capital_group_assets.append({
                **asset,
                "return": assumption["return"],
                "volatility": assumption["volatility"],
                "sourceMapping": assumption["sourceMapping"],
            })
        else:
            raw_capital_group_assets.append(None)
    invesco_assumptions = load_invesco_assumptions()
    raw_invesco_assets = []
    for asset in jpm_assets:
        normalized_asset_name = optimizer.normalize_name(asset["name"])
        if normalized_asset_name in invesco_assumptions:
            assumption = invesco_assumptions[normalized_asset_name]
            raw_invesco_assets.append({
                **asset,
                "return": assumption["return"],
                "volatility": assumption["volatility"],
                "sourceMapping": assumption["sourceMapping"],
            })
        else:
            raw_invesco_assets.append(None)
    msci_assumptions = load_msci_assumptions()
    raw_msci_assets = []
    for asset in jpm_assets:
        normalized_asset_name = optimizer.normalize_name(asset["name"])
        if normalized_asset_name in msci_assumptions:
            assumption = msci_assumptions[normalized_asset_name]
            raw_msci_assets.append({
                **asset,
                "return": assumption["return"],
                "volatility": assumption["volatility"],
                "sourceMapping": assumption["sourceMapping"],
            })
        else:
            raw_msci_assets.append(None)
    asset_allocation_interactive_assumptions = load_asset_allocation_interactive_assumptions()
    raw_asset_allocation_interactive_assets = []
    for asset in jpm_assets:
        normalized_asset_name = optimizer.normalize_name(asset["name"])
        if normalized_asset_name in asset_allocation_interactive_assumptions:
            assumption = asset_allocation_interactive_assumptions[normalized_asset_name]
            raw_asset_allocation_interactive_assets.append({
                **asset,
                "return": assumption["return"],
                "volatility": assumption["volatility"],
                "sourceMapping": assumption["sourceMapping"],
            })
        else:
            raw_asset_allocation_interactive_assets.append(None)
    average_assets = average_assumption_sets(
        [jpm_assets, raw_vanguard_assets, raw_blackrock_assets, raw_invesco_assets, raw_msci_assets, raw_capital_group_assets, raw_asset_allocation_interactive_assets],
        [JPM_FALLBACK_TO_AVERAGE, set(), BLACKROCK_FALLBACK_TO_AVERAGE, INVESCO_FALLBACK_TO_AVERAGE, set(), CAPITAL_GROUP_FALLBACK_TO_AVERAGE, ASSET_ALLOCATION_INTERACTIVE_FALLBACK_TO_AVERAGE],
    )
    jpm_assets = apply_average_fallback(jpm_assets, average_assets, "JPM 2026", JPM_FALLBACK_TO_AVERAGE)
    vanguard_assets = apply_average_fallback(raw_vanguard_assets, average_assets, "Vanguard 2026")
    blackrock_assets = apply_average_fallback(raw_blackrock_assets, average_assets, "BlackRock 2026", BLACKROCK_FALLBACK_TO_AVERAGE)
    invesco_assets = apply_average_fallback(raw_invesco_assets, average_assets, "Invesco 2026", INVESCO_FALLBACK_TO_AVERAGE)
    msci_assets = apply_average_fallback(raw_msci_assets, average_assets, "MSCI 2026")
    capital_group_assets = apply_average_fallback(raw_capital_group_assets, average_assets, "Capital Group 2026", CAPITAL_GROUP_FALLBACK_TO_AVERAGE)
    asset_allocation_interactive_assets = apply_average_fallback(raw_asset_allocation_interactive_assets, average_assets, "Asset Allocation Interactive 2026", ASSET_ALLOCATION_INTERACTIVE_FALLBACK_TO_AVERAGE)
    corr = blend_available_correlations(corr, average_assets)
    corr = nearest_valid_correlation(corr)
    large_cap = next(asset for asset in average_assets if asset["name"] == "U.S. Large Cap")
    jpm_returns: dict[str, float] = jpm_data["returns"]  # type: ignore[assignment]
    jpm_vols: dict[str, float] = jpm_data["vols"]  # type: ignore[assignment]
    benchmarks = [
        {
            "name": "S&P 500",
            "return": large_cap["return"],
            "volatility": large_cap["volatility"],
            "sourceMapping": "CORE U.S. Large Cap assumption",
        },
        {
            "name": "AGG",
            "return": jpm_returns[optimizer.normalize_name("U.S. Aggregate Bonds")],
            "volatility": jpm_vols[optimizer.normalize_name("U.S. Aggregate Bonds")],
            "sourceMapping": "JPM U.S. Aggregate Bonds assumption",
        },
    ]

    payload = {
        "generatedFrom": str(optimizer.jpm_matrix_path()),
        "cashTarget": optimizer.CASH_TARGET,
        "categories": optimizer.CATEGORIES,
        "selectedAssumptionSet": "CORE",
        "assets": average_assets,
        "assumptionSets": {
            "CORE": average_assets,
            "JPM 2026": jpm_assets,
            "Vanguard 2026": vanguard_assets,
            "BlackRock 2026": blackrock_assets,
            "Invesco 2026": invesco_assets,
            "MSCI 2026": msci_assets,
            "Capital Group 2026": capital_group_assets,
            "Asset Allocation Interactive 2026": asset_allocation_interactive_assets,
        },
        "volatilityModel": volatility_model,
        "benchmarks": benchmarks,
        "profiles": {
            profile: {
                "targetVolMin": config["target_volatility"][0],
                "targetVolMax": config["target_volatility"][1],
                "categoryBounds": {
                    category: {"min": bounds[0], "max": bounds[1]}
                    for category, bounds in config["category_bounds"].items()
                },
            }
            for profile, config in optimizer.PROFILES.items()
        },
        "correlation": corr.tolist(),
    }

    WEB_DATA.parent.mkdir(parents=True, exist_ok=True)
    WEB_DATA.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    print(WEB_DATA)


if __name__ == "__main__":
    main()
